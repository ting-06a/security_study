from openpyxl import load_workbook

workbook=load_workbook('read1.xlsx')

sheet = workbook.active

for row in sheet.iter_rows(min_row=1,min_col=1,max_row=10,max_col=2,values_only=True):
    print(row[1])